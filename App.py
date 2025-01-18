from re import sub
import sys, json
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QPushButton,
    QListView,
    QHBoxLayout,
    QTableView,
    QLabel,
    QDialog,
    QFormLayout,
    QLineEdit,
    QHeaderView,
    QComboBox,
    QFileDialog,
)
from PySide6.QtCore import (
    Qt,
    QAbstractListModel,
    QPoint,
    QAbstractTableModel,
    QModelIndex,
    QTimer,
    QDateTime
)
from PySide6.QtGui import QFont, QImage, QPainter, QColor, QBrush, QPixmap, QIcon, QAction, QKeySequence, QShortcut
from sqlalchemy import over
import db
import search
import requests
from openpyxl import Workbook
from datetime import datetime

current_version = 'v1.0'
def get_latest_version():
    github_url = "https://api.github.com/repos/mhieu56/schedule_app/releases/latest"
    
    response = requests.get(github_url)
    if response.status_code == 200:
        latest_release = response.json()
        latest_version = latest_release["tag_name"]  # Lấy tag version từ release mới nhất
        return latest_version
    else:
        print("Không thể lấy thông tin từ GitHub")
        return None

def get_latest_release():
    github_url = "https://api.github.com/repos/mhieu56/schedule_app/releases/latest"
    response = requests.get(github_url)
    if response.status_code == 200:
        latest_release = response.json()
        return latest_release
    else:
        print("Không thể lấy thông tin từ GitHub")
        return None

def get_asset_url(latest_release):
    # Tìm URL tải file .exe từ release
    assets = latest_release["assets"]
    for asset in assets:
        if asset["name"].endswith(".exe"):
            return asset["browser_download_url"]
    print("Không tìm thấy file .exe trong release.")
    return None

def download_file(url, filename):
    # Tải file về và lưu
    response = requests.get(url)
    if response.status_code == 200:
        with open(filename, "wb") as f:
            f.write(response.content)
        return True
    else:
        return False
        
def update_software():
    latest_release = get_latest_release()
    if latest_release:
        # Lấy URL tải về của file .exe
        asset_url = get_asset_url(latest_release)
        if asset_url:
            # Đặt tên file mới là v1.0.exe (hoặc tên tương ứng)
            filename = f"{get_latest_version()}.exe"
            
            dialog_1 = custom_dialog_3(title='Đang tải...')
            dialog_1.show()

            # Tải bản cập nhật mới
            if download_file(asset_url, filename):
                dialog_1.close()
                dialog_2 = custom_dialog_2(text=f'Đã tải xong {get_latest_version()}!',bold=False)
                dialog_2.exec()
            else:
                dialog_2 = custom_dialog_2(text='Tải thất bại!',bold=False)
                dialog_2.exec()
           
# Tạo dict để hiển thị trên TableWidget
subjects_list = db.get_subjects_name()
class_info_dict = {subjects_list[i]: None for i in range(len(subjects_list))}
button_height = 32
button_min_width = 150
button_max_width = 300

bold_font = QFont()
bold_font.setBold(True)
italic_font = QFont()
italic_font.setItalic(True)

#Chuyển đổi giữa số và chữ của ngày trong tuần
weekday_int = {
    "Thứ Hai": 2,
    "Thứ Ba": 3,
    "Thứ Tư": 4,
    "Thứ Năm": 5,
    "Thứ Sáu": 6,
    "Thứ Bảy": 7,
    "Chủ Nhật": 8,
}
int_weekday = {
    2: "Thứ Hai",
    3: "Thứ Ba",
    4: "Thứ Tư",
    5: "Thứ Năm",
    6: "Thứ Sáu",
    7: "Thứ Bảy",
    8: "Chủ Nhật",
}
#True sẽ hiển thị Login Form, False sẽ hiển thị thẳng vào ScheduleWindow
login = False
#Lưu danh sách những môn đang bị trùng lịch, 1: trùng, 0: không trùng
overlap_dict = {subjects_list[i]: 0 for i in range(len(subjects_list))}
#Danh sách key:value của môn học:id môn học
subject_id_dict = {subject[1]:subject[0] for subject in db.get_subjects_id_name()}

#Nút màu đỏ
red_button_style_string = """
    QPushButton {
            background-color:rgb(231, 59, 108);
            color: white;
            border-radius: 6px;
            outline: none;
            font-weight: bold;
            font-family: -apple-system, BlinkMacSystemFont, 'Roboto', sans-serif;
            border-bottom: 3px solid rgb(184, 60, 95);
        }
        QPushButton:hover {
            background-color: rgb(206, 72, 110);
            outline: None;
        }
"""


class TableModel(QAbstractTableModel):

    def __init__(self, data, headers):
        super().__init__()
        self._data = data
        self._headers = headers

    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if index.column() == 0:
                return None
            if index.column() == 5:
                weekday = self._data[index.row()][index.column()]
                return int_weekday[weekday]
            return self._data[index.row()][index.column()]
        if role == Qt.TextAlignmentRole:
            header = self._headers[index.column()]
            if header == "Tên môn" or header == "Giảng viên":
                return Qt.AlignLeft | Qt.AlignVCenter
            else:
                return Qt.AlignCenter

    def rowCount(self, index=QModelIndex()):
        return len(self._data)

    def columnCount(self, index=QModelIndex()):
        if self._data and len(self._data) > 0:
            return len(self._data[0])
        return 0

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._headers[section]
            elif orientation == Qt.Vertical:
                return section + 1
        return None


# Hàm tạo Dialog với text=Truyền vào
class custom_dialog(QDialog):
    def __init__(self, parent=None, text="Bạn chắc chứ?", bold=True):
        super().__init__()

        self.setWindowTitle("Second Thought?")
        self.setModal(True)
        self.setMinimumWidth(250)

        main_layout = QVBoxLayout()
        self.label = QLabel()
        self.label.setText(text)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("padding: 10px; font-size: 15px; color: #474a4d;")
        if bold:
            self.label.setFont(bold_font)

        button_layout = QHBoxLayout()
        self.ok_button = QPushButton(text="Chắc Chắn")
        self.ok_button.pressed.connect(self.accept)
        self.ok_button.setMinimumHeight(button_height)
        self.ok_button.setMinimumWidth(120)
        self.ok_button.setMaximumWidth(120)
        self.ok_button.setStyleSheet(red_button_style_string)
        self.cancel_button = QPushButton(text="Thôi khỏi")
        self.cancel_button.pressed.connect(self.reject)
        self.cancel_button.setMinimumHeight(button_height)
        self.cancel_button.setMinimumWidth(120)
        self.cancel_button.setMaximumWidth(120)
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        button_layout.setAlignment(Qt.AlignCenter)

        main_layout.addWidget(self.label)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)


#Update Version Dialog
class update_version_dialog(QDialog):
    def __init__(self, parent=None, text="Bạn chắc chứ?"):
        super().__init__()

        self.setWindowTitle("Tìm kiếm bản cập nhật")
        self.setModal(True)
        self.setMinimumWidth(250)

        main_layout = QVBoxLayout()
        self.label = QLabel()
        self.label.setText(text)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("padding: 10px; font-size: 15px; color: #474a4d;")

        button_layout = QHBoxLayout()
        self.ok_button = QPushButton(text="Chắc Chắn")
        self.ok_button.pressed.connect(self.accept)
        self.ok_button.setMinimumHeight(button_height)
        self.ok_button.setMinimumWidth(120)
        self.ok_button.setMaximumWidth(120)
        self.cancel_button = QPushButton(text="Thôi khỏi")
        self.cancel_button.pressed.connect(self.reject)
        self.cancel_button.setMinimumHeight(button_height)
        self.cancel_button.setMinimumWidth(120)
        self.cancel_button.setMaximumWidth(120)
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        button_layout.setAlignment(Qt.AlignCenter)

        main_layout.addWidget(self.label)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)


class CountdownApp(QDialog):
    def __init__(self,parent=None,future=QDateTime(2025, 2, 6, 8, 0, 0)):
        super().__init__()
        
        current_time = QDateTime.currentDateTime()

        # Khởi tạo ngày giờ trong tương lai (ví dụ: 2025-01-20 10:30)
        self.future_datetime = future
        
        time_left = current_time.secsTo(self.future_datetime)

        if time_left > 0:
            days_left = time_left // 86400  # Số ngày còn lại
            hours_left = (time_left % 86400) // 3600  # Số giờ còn lại
            minutes_left = (time_left % 3600) // 60  # Số phút còn lại
            seconds_left = time_left % 60  # Số giây còn lại

        # Tạo các QLabel để hiển thị thời gian hiện tại, thời gian trong tương lai và thời gian còn lại
        self.current_time_label = QLabel("Hiện tại là:")
        self.current_time_label.setAlignment(Qt.AlignCenter)
        self.current_time_label_2 = QLabel(current_time.toString('HH:mm:ss   dd-MM-yyyy'))
        self.future_time_label = QLabel("Ngày đăng ký học phần:")
        self.future_time_label.setAlignment(Qt.AlignCenter)
        self.future_time_label_2 = QLabel(self.future_datetime.toString('HH:mm:ss   dd-MM-yyyy'))
        self.time_left_label = QLabel("Bạn còn: ")
        self.time_left_label.setAlignment(Qt.AlignCenter)
        self.time_left_label_2 = QLabel("")
        self.time_left_label_2.setText(f"{days_left} Ngày {hours_left:02}:{minutes_left:02}:{seconds_left:02}")
        
        #Thiết lập style cho các label
        label_list = [self.current_time_label,self.future_time_label,self.time_left_label]
        label_time_list = [self.current_time_label_2,self.future_time_label_2,self.time_left_label_2]
        for label in label_list + label_time_list:
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("padding: 0px;")
        for label in label_list:
            label.setStyleSheet("font-size: 15px; color: #424242;")
        for label in label_time_list:
            label.setStyleSheet("font-size: 20px; color: #424242; font-weight: bold;")
        self.future_time_label.setStyleSheet("font-size: 15px; color: #424242;padding-top: 15px;")

        # Thiết lập layout
        layout = QVBoxLayout()
        layout.addWidget(self.current_time_label)
        layout.addWidget(self.current_time_label_2)
        layout.addWidget(self.future_time_label)
        layout.addWidget(self.future_time_label_2)
        layout.addWidget(QLabel())
        layout.addWidget(self.time_left_label)
        layout.addWidget(self.time_left_label_2)
        
        self.ok_button = QPushButton(text="Xác Nhận")
        self.ok_button.pressed.connect(self.accept)
        self.ok_button.setMinimumWidth(150)
        self.ok_button.setMinimumHeight(35)

        layout.addWidget(QLabel())
        layout.addWidget(self.ok_button, alignment=Qt.AlignCenter)
        
        self.setLayout(layout)

        # Khởi tạo QTimer để cập nhật thời gian mỗi giây
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time)
        self.timer.start(1000)  # Cập nhật mỗi giây

        # Thiết lập cửa sổ
        self.setWindowTitle("Countdown Timer")
        self.resize(300,300)

    def update_time(self):
        # Lấy thời gian hiện tại
        current_time = QDateTime.currentDateTime()
        self.current_time_label_2.setText(current_time.toString('HH:mm:ss   dd-MM-yyyy'))

        # Tính toán thời gian còn lại
        time_left = current_time.secsTo(self.future_datetime)

        if time_left > 0:
            days_left = time_left // 86400  # Số ngày còn lại
            hours_left = (time_left % 86400) // 3600  # Số giờ còn lại
            minutes_left = (time_left % 3600) // 60  # Số phút còn lại
            seconds_left = time_left % 60  # Số giây còn lại

            self.time_left_label_2.setText(f"{days_left} Ngày {hours_left:02}:{minutes_left:02}:{seconds_left:02}")
        else:
            self.time_left_label.setText("Time Left: Time's up!")


# Tạo Dialog chỉ có nút "Xác Nhận"
class custom_dialog_2(QDialog):
    def __init__(self, parent=None, text="Chắc hông?", bold=True):
        super().__init__()
        self.setWindowTitle("Second Thought")
        self.setModal(True)
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        self.label = QLabel(text)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("padding: 10px; font-size: 16px; color: #474a4d;")
        if bold:
            self.label.setFont(bold_font)
        layout.addWidget(self.label)

        self.ok_button = QPushButton(text="Xác Nhận")
        self.ok_button.pressed.connect(self.accept)
        self.ok_button.setMinimumWidth(150)
        self.ok_button.setMinimumHeight(35)

        layout.addWidget(self.ok_button, alignment=Qt.AlignCenter)
        self.setLayout(layout)
        self.resize(250, 120)


# Tạo Dialog không có nút gì hết
class custom_dialog_3(QDialog):
    def __init__(self, parent=None,title='Second Thought', text="Chắc hông?"):
        super().__init__()
        self.setWindowTitle(title)
        self.setModal(True)
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        self.label = QLabel(text)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("padding: 10px; font-size: 16px; color: #474a4d;")
        layout.addWidget(self.label)

        self.setLayout(layout)
        self.resize(250, 120)


class add_subject_to_database_dialog(QDialog):
    def __init__(self, parent=None, subject_list=[]):
        super().__init__()
        self.setWindowTitle("Second Thought")
        self.setModal(True)
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        subject_string = ''
        for i in subject_list:
            subject_string += f'{i}\n'
        
        self.label_1 = QLabel('Cơ sở dữ liệu hiện tại của bạn thiếu những môn sau:\n')
        self.label_1.setAlignment(Qt.AlignCenter)
        self.subject_list_label = QLabel(subject_string)
        self.subject_list_label.setFont(bold_font)
        self.subject_list_label.setAlignment(Qt.AlignCenter)
        self.label_2 = QLabel('Muốn thêm vào hông?\n')
        self.label_2.setAlignment(Qt.AlignCenter)

        layout.addWidget(self.label_1)
        layout.addWidget(self.subject_list_label)
        layout.addWidget(self.label_2)
        
        button_layout = QHBoxLayout()

        self.ok_button = QPushButton(text="Đồng Ý")
        self.ok_button.pressed.connect(self.accept)
        self.ok_button.setMinimumWidth(150)
        self.ok_button.setMaximumWidth(150)
        self.ok_button.setMinimumHeight(35)
        self.ok_button.setFont(bold_font)
        
        self.cancel_button = QPushButton(text="Thôi Khỏi")
        self.cancel_button.pressed.connect(self.reject)
        self.cancel_button.setMinimumWidth(150)
        self.cancel_button.setMaximumWidth(150)
        self.cancel_button.setMinimumHeight(35)
        
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)
        self.resize(250, 120)


#Subject Model
class ListModel(QAbstractListModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data or []

    def create_pixmap_from_int(self, value, extra_text=""):
        pixmap = QPixmap(35, 20)  # Kích thước hình ảnh
        pixmap.fill(Qt.transparent)  # Nền trong suốt
        painter = QPainter(pixmap)
        painter.setPen(QColor("#054223"))  # Màu chữ
        painter.drawText(
            pixmap.rect(),
            Qt.AlignVCenter | Qt.AlignRight,
            str(f"{extra_text} {value:02}"),
        )  # Vẽ số vào hình ảnh
        painter.end()
        return pixmap

    def data(self, index, role):
        text = self._data[index.row()]
        if role == Qt.DisplayRole:
            return text
        if role == Qt.DecorationRole:
            value = db.class_count(text)
            if class_info_dict[text]:
                if overlap_dict.get(text) == 1:
                    return self.create_pixmap_from_int(value, extra_text="❌")
                else:
                    return self.create_pixmap_from_int(value, extra_text="✔️")
            else:
                return self.create_pixmap_from_int(value)
        if role == Qt.ForegroundRole:
            # Lấy giá trị từ overlap_dict trực tiếp
            if overlap_dict[text]:
                return QColor("#e3143e")
        if role == Qt.FontRole:
            if overlap_dict[text]:
                return bold_font

    def rowCount(self, index):
        if not self._data:
            return 0
        return len(self._data)
    
    def refresh(self, new_data):
        """Làm mới lại model."""
        self.beginResetModel()
        self._data = new_data
        self.endResetModel()


class list_model(QAbstractListModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = data or []

    
    def data(self, index, role):
        text = self._data[index.row()]
        if role == Qt.DisplayRole:
            return text

    def rowCount(self, index):
        if not self._data:
            return 0
        return len(self._data)


class find_subject_form(QDialog):
    def __init__(self, parent=None, subject_list_model=ListModel()):
        super().__init__(parent)

        button_height = 35
        button_min_width = 150
        button_max_width = 300

        self.setWindowTitle("Form Tra Cứu")
        self.setGeometry(150, 120, 1200, 600)
        self.setModal(True)

        form_layout = QFormLayout()

        self.subject_id_edit = QLineEdit()
        self.subject_id_edit.setMaximumWidth(250)
        self.subject_id_edit.setStyleSheet("padding: 5px 5px;")
        center_layout = QHBoxLayout()
        center_layout.addStretch(1)
        center_layout.addWidget(QLabel("Mã môn:  "))
        center_layout.addWidget(self.subject_id_edit, stretch=2)
        #Nút tra cứu
        self.search_button = QPushButton(text="Tra Cứu")
        self.search_button.setMinimumHeight(button_height-6)
        self.search_button.setMinimumWidth(button_min_width-40)
        self.search_button.setMaximumWidth(button_max_width-40)
        self.search_button.clicked.connect(self.search_class)
        self.search_button.setShortcut(QKeySequence("Enter"))
        center_layout.addWidget(self.search_button)
        self.subject_name = QLabel("")
        self.subject_name.setAlignment(Qt.AlignLeft)
        center_layout.addWidget(self.subject_name, stretch=1)
        form_layout.addRow(center_layout)
        self.note_label = QLabel("Danh sách lớp học bên dưới lấy từ cơ sở dữ liệu của Trường Cao đẳng Công nghệ Thủ Đức")
        self.note_label.setAlignment(Qt.AlignCenter)
        self.note_label.setStyleSheet("""
            QLabel {
                font-style: italic;
                color: #666666;
                padding: 4px;
            }
                                      """)
        form_layout.addRow(self.note_label)
        #Layout list môn hiện tại và tất cả môn học
        self.subject_list_layout = QHBoxLayout()
        current_subject_layout = QHBoxLayout()
        self.current_subject_list = QComboBox()
        self.current_subject_list.currentIndexChanged.connect(self.on_current_subject_change)
        self.current_subject_list.setStyleSheet("min-width: 150px; max-width: 150px; padding: 5px;")
        self.current_subject_list.addItems(db.get_subjects_name())
        self.current_subject_list.setCurrentIndex(-1)
        current_subject_layout.addWidget(QLabel("Môn hiện tại:"))
        current_subject_layout.addWidget(self.current_subject_list)
        current_subject_layout.setAlignment(Qt.AlignCenter)
        self.subject_list_layout.addStretch(1)
        self.subject_list_layout.addLayout(current_subject_layout)
        self.subject_list_layout.addWidget(QLabel("    "))
        #Tìm kiếm môn học
        all_subject_layout = QHBoxLayout()
        self.find_subject = QLineEdit()
        self.find_subject.setStyleSheet("min-width: 150px; max-width: 150px; padding: 5px;")
        self.find_subject.setPlaceholderText("Nhập tên môn học..")
        self.find_subject.textChanged.connect(self.search_class_by_name)
        all_subject_layout.addWidget(QLabel("Tìm kiếm:"))
        all_subject_layout.addWidget(self.find_subject)
        all_subject_layout.setAlignment(Qt.AlignCenter)
        self.subject_list_layout.addLayout(all_subject_layout)
        self.subject_list_layout.addWidget(QLabel("    "))
        #Danh sách môn học theo khoa
        subjetcs_by_faculty_layout = QHBoxLayout()
        subjetcs_by_faculty_layout.addWidget(QLabel("Khoa:"))
        self.faculty = QComboBox()
        self.faculty.setMaximumWidth(250)
        self.faculty.setStyleSheet("padding: 5px;")
        self.faculty.addItems(db.get_all_faculty())
        self.faculty.setCurrentIndex(-1)
        self.faculty.currentIndexChanged.connect(self.on_faculty_change)
        subjetcs_by_faculty_layout.addWidget(self.faculty)
        subjetcs_by_faculty_layout.addWidget(QLabel("Học kỳ:"))
        self.semester = QComboBox()
        self.semester.setStyleSheet("min-width: 70px; max-width: 70px; padding: 5px;")
        self.semester.currentIndexChanged.connect(self.on_semester_change)
        subjetcs_by_faculty_layout.addWidget(self.semester)
        self.subjetcs_by_faculty_list = QComboBox()
        self.subjetcs_by_faculty_list.setStyleSheet("padding: 5px;")
        self.subjetcs_by_faculty_list.setMaximumWidth(200)
        self.subjetcs_by_faculty_list.setMinimumWidth(200)
        self.subjetcs_by_faculty_list.currentIndexChanged.connect(self.on_subjects_by_change)
        subjetcs_by_faculty_layout.addWidget(self.subjetcs_by_faculty_list)
        self.subject_list_layout.addLayout(subjetcs_by_faculty_layout)
        self.subject_list_layout.addStretch(1)
        
        form_layout.addRow(self.subject_list_layout)
        
        headers = [
            "Id",
            "Mã môn",
            "Mã lớp",
            "Tên lớp",
            "Giảng viên",
            "Thứ",
            "Bắt đầu",
            "Kết thúc",
            "Phòng",
            "Ngày bắt đầu",
            "Ngày kết thúc",
        ]
        self.classes_filter = QTableView()
        self.classes_list = db.get_classes("")
        self.classes_model = TableModel(self.classes_list, headers)
        self.classes_filter.setModel(self.classes_model)
        self.classes_filter.hideColumn(0)
        self.classes_filter.setColumnWidth(4,200)
        

        #Hàng nút ở dưới
        button_layout = QHBoxLayout()
        button_style_string = "padding: 5px 12px;"
        self.export_excel_button = QPushButton("Xuất File Excel")
        self.export_excel_button.clicked.connect(self.export_excel)
        self.export_excel_button.setStyleSheet(button_style_string)
        self.update_button = QPushButton("Cập Nhật Vào DB")
        self.update_button.clicked.connect(self.update_database)
        self.update_button.setStyleSheet(button_style_string)
        self.update_faculty_button = QPushButton(f'Cập Nhật theo Khoa + Học kỳ')
        self.update_faculty_button.clicked.connect(self.update_faculty_database)
        self.update_faculty_button.setStyleSheet(button_style_string)
        self.filter_button = QPushButton("Lọc Đã Qua")
        self.filter_button.clicked.connect(self.filter_over)
        self.filter_button.setStyleSheet(button_style_string)
        self.filter_night_button = QPushButton("Lọc Lớp Tối")
        self.filter_night_button.clicked.connect(self.filter_night)
        self.filter_night_button.setStyleSheet(button_style_string)
        button_layout.addWidget(self.update_button)
        button_layout.addWidget(QLabel("    "))
        button_layout.addWidget(self.filter_button)
        button_layout.addWidget(self.filter_night_button)
        button_layout.addWidget(QLabel("    "))
        button_layout.addWidget(self.export_excel_button)
        button_layout.setAlignment(Qt.AlignCenter)

        button_h_layout = QHBoxLayout()
        button_h_layout.addStretch(1)
        button_h_layout.addLayout(button_layout)
        button_h_layout.addStretch(1)
        
        form_layout.addRow(QLabel(''))
        form_layout.addRow(button_h_layout)
        form_layout.addRow(self.classes_filter)

        self.setLayout(form_layout)
    
    
    def update_faculty_database(self):
        if self.faculty.currentIndex() != -1:
            faculty_name = self.faculty.currentText()
            semester = self.semester.currentText()
            subject_name_list = db.get_subjects_by_faculty_semester(faculty_name,semester)
            
            self.add_subject_to_database(subject_name_list)
            self.update_current_subject()
        else:
            custom_dialog_2(self,"Vui lòng chọn Khoa và Học kỳ!",False)
                     
    def add_subject_to_database(self, subject_list=None):
        if subject_list:
            for subject_name in subject_list:
                subject_id = db.get_subject_id_by_name(subject_name)
                if subject_id is not None:  # Kiểm tra nếu subject_id hợp lệ
                    subject_info = (subject_id, subject_name)
                    db.add_subject(subject_info)
                else:
                    print(f"Subject '{subject_name}' not found in database!")  
    
    def on_subjects_by_change(self):
        faculty_name = self.faculty.currentText()
        subject_name = self.subjetcs_by_faculty_list.currentText()
        if faculty_name and subject_name:
            subject_id = db.get_subject_id_by_faculty_subject_name(faculty_name,subject_name)
            self.search_class(subject_id)
    
    def on_semester_change(self):
        faculty_name = self.faculty.currentText()
        semester = self.semester.currentText()
        subjects_list = self.subjetcs_by_faculty_list
        if faculty_name and semester:
            subjects_list.clear()
            subjects_list.addItems(db.get_subjects_by_faculty_semester(faculty_name,semester))
        else:
            subjects_list.clear()
    
    def on_faculty_change(self):
        semester_list = self.semester
        faculty_name = self.faculty.currentText()
        if faculty_name:
            semester_list.clear()
            semester_list.addItems(db.get_semester_by_faculty(faculty_name))
        else:
            semester_list.clear()
    
    def search_class_by_name(self):
        text = self.find_subject.text().strip()
        if text:
            data = db.find_subject(text)
            self.classes_model = TableModel(data, ['id','Mã môn','Tên môn'])
            self.classes_filter.setModel(self.classes_model)
            self.classes_filter.hideColumn(0)
            self.classes_filter.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def on_current_subject_change(self):
        if self.current_subject_list.currentIndex() == -1:
            return
        suject_id = subject_id_dict[self.current_subject_list.currentText()]
        self.search_class(suject_id)

    def filter_night(self,_data=None):
        if not _data:
            data = self.classes_model._data
        else:
            data = _data
        time_2 = datetime.strptime("18:00", "%H:%M").time()
        if data:
                data_filtered = [class_info for class_info in data if datetime.strptime(class_info[6], "%H:%M").time() < time_2]
        if not _data:
            self.classes_model._data = data_filtered
            self.classes_model.layoutChanged.emit()
            self.classes_filter.hideColumn(0)
            self.classes_filter.setColumnWidth(3,150)
            self.classes_filter.setColumnWidth(4,200)
            self.classes_filter.setColumnWidth(5,80)
        else:
            return data_filtered
            
    def filter_over(self,_data=None):
        if not _data:
            data = self.classes_model._data
        else:
            data = _data
        today = datetime.today().date()
        if data:
            data_filtered = [class_info for class_info in data if datetime.strptime(class_info[9], "%d/%m/%Y").date()>today]
        if not _data:
            self.classes_model._data = data_filtered
            self.classes_model.layoutChanged.emit()
            self.classes_filter.hideColumn(0)
            self.classes_filter.setColumnWidth(3,150)
            self.classes_filter.setColumnWidth(4,200)
            self.classes_filter.setColumnWidth(5,80)
        else:
            return data_filtered
        
    def search_class(self,subject_id=None):
        if not subject_id:
            data = search.search(self.subject_id_edit.text().strip())
        else:
            data = search.search(subject_id.strip())
        if data:
            headers = [
                "Id",
                "Mã môn",
                "Mã lớp",
                "Tên lớp",
                "Giảng viên",
                "Thứ",
                "Bắt đầu",
                "Kết thúc",
                "Phòng",
                "Ngày bắt đầu",
                "Ngày kết thúc",
            ]
            self.classes_model._headers = headers
            self.classes_model._data = data
            self.classes_filter.setSelectionBehavior(QTableView.SelectRows)
            self.classes_model.layoutChanged.emit()
            self.subject_name.setText("Môn:  " + data[0][3].upper())
            self.subject_name.setAlignment(Qt.AlignCenter)
            self.subject_name.setFont(bold_font)
            self.classes_filter.hideColumn(0)
            self.classes_filter.setColumnWidth(3,150)
            self.classes_filter.setColumnWidth(4,200)
            self.classes_filter.setColumnWidth(5,80)
        else:
            self.classes_model._headers = []
            self.classes_model._data = []
            self.classes_filter.setSelectionBehavior(QTableView.SelectRows)
            self.classes_model.layoutChanged.emit()
            self.subject_name.setText('')
                
    def export_excel(self):
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,  # Widget cha
            "Lưu tệp",  # Tiêu đề hộp thoại
            "Output",  # Đường dẫn mặc định
            "Excel Files (*.xlsx)",  # Bộ lọc định dạng tệp
        )
        
        model = self.classes_model
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Ghi tiêu đề cột
        for column in range(model.columnCount()):
            header = model.headerData(column, orientation=Qt.Horizontal)  # orientation=1 là horizontal (tiêu đề cột)
            sheet.cell(row=1, column=column + 1, value=header)
            
        # Ghi dữ liệu hàng
        for row in range(model.rowCount()):
            for column in range(model.columnCount()):
                value = model.data(model.index(row, column))
                sheet.cell(row=row + 2, column=column + 1, value=value)
                
        # Lưu file
        if file_path:
            workbook.save(file_path)
            dialog = custom_dialog_2(None, f"Đã lưu file vào:   {file_path}", False)
            dialog.exec()
    
    def update_database(self,subject_list_model):
        data = self.classes_model._data
        class_list = []
        if data:
            subject_id = data[0][1]
            subject_name = data[0][3]
            for row in data:
                class_row = (
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    row[6],
                    row[7],
                    row[8]
                )
                class_list.append(class_row)
            db.update_database(subject_id,subject_name,class_list)
            dialog = custom_dialog_2(self,"Cập nhật thành công!", False)
            dialog.exec()
        

class subject_by_facaulty_dialog(QDialog):
    def __init__(self, parent=None):
        super().__init__()
        self.setWindowTitle("Form Chọn Khoa và Học kỳ")
        self.setModal(True)
        
        form_layout = QFormLayout()
        
        self.faculty = QComboBox()
        self.faculty.addItems(db.get_all_faculty())
        self.faculty.currentIndexChanged.connect(self.on_faculty_change)
        self.faculty.setStyleSheet("padding: 5px 10px;")
        self.faculty.setCurrentIndex(-1)
        form_layout.addRow('Khoa: ',self.faculty)
        
        self.semester = QComboBox()
        self.semester.setStyleSheet("padding: 5px 10px;")
        self.semester.currentIndexChanged.connect(self.on_semester_change)
        form_layout.addRow('Học kỳ:',self.semester)

        button_layout = QHBoxLayout()
        button_layout.setAlignment(Qt.AlignCenter)
        self.ok_button = QPushButton('Xác Nhận')
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setMinimumWidth(150)
        self.ok_button.setMinimumHeight(35)
        
        self.cancel_button = QPushButton('Hủy')
        self.cancel_button.clicked.connect(self.reject)
        self.cancel_button.setMinimumWidth(150)
        self.cancel_button.setMinimumHeight(35)

        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        
        self.class_list_view = QListView()
        form_layout.addRow('',self.class_list_view)
        
        form_layout.addRow('', QLabel())
        form_layout.addRow('',button_layout)
        self.setLayout(form_layout)
    
    def get_semester(self):
        return self.semester.currentText()
    
    def get_faculty_name(self):
        return self.faculty.currentText()
    
    def on_semester_change(self):
        faculty = self.faculty.currentText()
        semester = self.semester.currentText()
        data = db.get_subjects_by_faculty_semester(faculty,semester)
        class_view_model = list_model(data)
        self.class_list_view.setModel(class_view_model)
    
    def on_faculty_change(self):
        faculty_name = self.faculty.currentText()
        self.semester.clear()
        self.semester.addItems(db.get_semester_by_faculty(faculty_name))


def export_to_image(table_view, file_name):
    table_view.repaint()
    # Determine the size of the table view
    table_width = sum(table_view.columnWidth(col) for col in range(table_view.columnCount()))
    table_height = table_view.verticalHeader().length() + table_view.horizontalHeader().height()

    # Create an image and render the table view onto it
    image = QImage(table_width+table_view.columnWidth(0), table_height+20, QImage.Format_ARGB32)
    image.fill(0xFFFFFF)  # Set white background
    painter = QPainter(image)
    table_view.render(painter, QPoint(0, 0))
    painter.end()

    # Save the image
    image.save(file_name)
    dialog = custom_dialog_2(None, f"Đã lưu ảnh vào:   {file_name}", False)
    dialog.exec()


class ScheduleWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Tạo Thời Khóa Biểu")
        self.setGeometry(100, 50, 1300, 750)

        self.disable_button_style_string = """
            QPushButton {
                color: #808080; 
                background-color: #d6d6d6; 
                font-weight: bold;
                border-radius: 6px;
                outline: none;
                font-family: -apple-system, BlinkMacSystemFont, 'Roboto', sans-serif;
                border-bottom: 3px solid rgb(135, 135, 135);
            }
        """
        
        self.default_font = QFont()
        self.default_font.setBold(False)

        #Tạo Menubar
        menubar = self.menuBar()
        menubar.setStyleSheet("""
            QMenuBar {
                font-size: 14px;
                font-family: Arial, sans-serif;
                border-bottom: 1px solid #e0e0e0;
                background-color: #ffffff;
            }
            QMenuBar::item {
                padding: 3px 6px;
                font-size: 10px;
            }
            QMenu {
                border: 1px solid #7a7a7a;
                border-top: none;
                border-radius: 0;
                background-color: #ffffff;
            }
            QMenu::item {
                padding: 5px 10px;            
            }
            QMenu::item:selected {
                color: #5c5c5c;
            }
            QMenu::separator {
                background: #f0f0f0;
                height: 1px;
            }
        """)
        
        space_string = '          '
        file_menu = menubar.addMenu('File')
        open_action = QAction(f'Mở file{space_string}',self)
        open_action.setShortcut(QKeySequence("Ctrl+O"))
        open_action.triggered.connect(self.open_file)
        file_menu.addAction(open_action)
        
        save_action = QAction('Lưu file',self)
        save_action.triggered.connect(self.save_file)
        save_action.setShortcut(QKeySequence("Ctrl+S"))
        file_menu.addAction(save_action)
        
        file_menu.addSeparator()

        exit_action = QAction('Thoát',self)
        exit_action.triggered.connect(self.close_app)
        exit_action.setShortcut(QKeySequence("Alt+F4"))
        file_menu.addAction(exit_action)
        
        refresh_shortcut = QShortcut(QKeySequence("F5"), self)
        refresh_shortcut.activated.connect(self.refresh_view)
        
        help_menu = menubar.addMenu('Help')
        
        countdown_action = QAction('Countdown',self)
        countdown_action.triggered.connect(self.countdown)
        update_action = QAction(f'Update{space_string}', self)
        update_action.triggered.connect(self.update_version)
        help_menu.addAction(countdown_action)
        help_menu.addAction(update_action)

        # Tạo widget chính
        widget = QWidget(self)
        self.setCentralWidget(widget)
        layout = QVBoxLayout(widget)

        # Tạo QTableWidget
        self.table = QTableWidget(self)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setRowCount(14)  # 2 buổi, mỗi buổi 6 mốc thời gian
        self.table.setColumnCount(7)  # 7 ngày trong tuần (từ thứ 2 đến chủ nhật)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # Đặt tiêu đề cột (thứ trong tuần)
        self.table.setHorizontalHeaderLabels(
            ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ Nhật"]
        )

        # Đặt tiêu đề hàng (sáng và chiều)
        self.table.setVerticalHeaderLabels(
            [
                "07:00",
                "07:45",
                "08:30 - 08:55",
                "09:40",
                "10:30",
                "11:15",
                "12:00",
                "12:45",
                "13:30",
                "14:15",
                "15:00 - 15:25",
                "16:10",
                "16:55",
                "17:40",
            ]
        )

        # Đặt nội dung cho bảng
        self.set_schedule_data()

        self.customize_header()

        layout.addWidget(self.table, stretch=6)

        # Label cho hai thằng danh sách Môn và danh sách Lớp
        subject_class_label_layout = QHBoxLayout()
        self.subject_label = QLabel()
        self.subject_label.setText("Môn học:   0 đã chọn")
        self.subject_label.setContentsMargins(10, 0, 0, 0)
        
        #Layout phần hiển thị lớp
        self.class_layout = QHBoxLayout()
        self.class_label = QLabel("Lớp học tương ứng:")
        self.class_label.setContentsMargins(10, 0, 0, 0)
        self.class_layout.addWidget(self.class_label)
        
        #Bộ công cụ QPushButton cho lớp
        self.class_button_layout = QHBoxLayout()
        self.filter_category = QComboBox()
        self.filter_category.addItems(["Tất cả", "Giảng viên", "Thứ", "Buổi"])
        self.filter_category.setMinimumWidth(120)
        self.filter_category.setMaximumWidth(120)
        self.filter_category.currentIndexChanged.connect(self.on_filter_category_change)
        self.filter_list = QComboBox()
        self.filter_list.setMinimumWidth(180)
        self.filter_list.setMaximumWidth(250)
        self.filter_list.currentIndexChanged.connect(self.on_filter_list_change)
        self.filter_label = QLabel("Lọc:")
        self.class_button_layout.addWidget(self.filter_label)
        self.class_button_layout.addWidget(self.filter_category)
        self.class_button_layout.addWidget(self.filter_list)
        self.filter_label.setVisible(False)
        self.filter_category.setVisible(False)
        self.filter_list.setVisible(False)
        self.class_layout.addLayout(self.class_button_layout)
        self.class_button_layout.setAlignment(Qt.AlignRight)
        subject_class_label_layout.addWidget(self.subject_label, stretch=1)
        subject_class_label_layout.addLayout(self.class_layout, stretch=5)

        subject_class_layout = QHBoxLayout()
        # List Môn Học
        self.subjects = QListView()
        self.subjects_list = db.get_subjects_name()
        self.subjects_model = ListModel(self.subjects_list)
        self.subjects.setModel(self.subjects_model)
        self.subjects.selectionModel().selectionChanged.connect(self.on_subject_view_change)

        # List Lớp Học tương ứng
        self.classes = QTableView()
        self.classes.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        headers = [
            "Id",
            "Mã môn",
            "Mã lớp",
            "Tên lớp",
            "Giảng viên",
            "Thứ",
            "Bắt đầu",
            "Kết thúc",
            "Phòng",
        ]
        self.classes_list = db.get_classes("Cơ sở dữ liệu")
        self.classes_model = TableModel(self.classes_list, headers)
        self.classes.setModel(self.classes_model)
        self.classes.hideColumn(0)
        self.classes.selectionModel().selectionChanged.connect(self.on_class_table_change)  # Cập nhật thời khóa biểu khi người dùng chọn lớp
        self.classes.setSelectionBehavior(QTableView.SelectRows)
        subject_class_layout.addWidget(self.subjects, stretch=1)
        subject_class_layout.addWidget(self.classes, stretch=5)

        # Thêm nút
        # Thêm nút để kiểm tra khả năng chỉnh sửa
        self.button_height = 30
        button_layout = QHBoxLayout()
        self.sort_button = QPushButton('Sắp Xếp')
        self.sort_button.clicked.connect(self.sort_subject_list)
        self.update_current_subject_button = QPushButton("Cập Nhật Nhanh")
        self.update_current_subject_button.clicked.connect(self.update_current_subject)
        self.update_current_subject_button.setToolTip("Cập nhật lớp của những môn hiện tại")
        self.update_by_faculty_button = QPushButton("Cập Nhật Theo Khoa")
        self.update_by_faculty_button.clicked.connect(self.update_by_faculty_subject)
        self.export_button = QPushButton("Lưu Ảnh", self)
        self.export_button.clicked.connect(self.export_schedule)
        self.add_subject_button = QPushButton("Tra Cứu", self)
        self.add_subject_button.clicked.connect(self.add_subject)
        self.delete_subject_button = QPushButton("Xóa Môn", self)
        self.delete_subject_button.clicked.connect(self.delete_subject)
        self.delete_all_button = QPushButton("Xóa Tất Cả", self)
        self.delete_all_button.clicked.connect(self.delete_all)
        self.deselect_button = QPushButton("Bỏ Chọn", self)
        self.deselect_button.clicked.connect(self.deselect)
        self.deselect_button.setStyleSheet(self.disable_button_style_string)
        
        #Danh sách nút
        button_normal_list = [self.sort_button,self.update_by_faculty_button, self.export_button,self.add_subject_button,self.deselect_button,self.update_current_subject_button]
        button_red_list = [self.delete_subject_button,self.delete_all_button]
        #Set style cho nút bình thường
        for button in button_normal_list + button_red_list:
            button.setMinimumHeight(button_height)
            button.setMaximumWidth(160)
        
        #Set style các cho nút xóa
        for button in button_red_list:
            button.setStyleSheet(red_button_style_string)
        
        self.sort_button.setMaximumWidth(90)
        self.update_current_subject_button.setMaximumWidth(130)
        
        button_layout.addWidget(self.add_subject_button)
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.deselect_button)
        button_layout.addWidget(QLabel(" "))
        button_layout.addWidget(self.delete_subject_button)
        button_layout.addWidget(self.delete_all_button)
        button_h_layout = QHBoxLayout()
        update_h_layout = QHBoxLayout()
        update_h_layout.addWidget(self.sort_button)
        update_h_layout.addWidget(self.update_current_subject_button)
        update_h_layout.addWidget(self.update_by_faculty_button)
        button_h_layout.addLayout(update_h_layout, stretch=2)
        button_h_layout.addStretch(1)
        button_h_layout.addLayout(button_layout, stretch=3)
        button_h_layout.addStretch(1)
        layout.addLayout(button_h_layout)

        layout.addLayout(subject_class_label_layout)
        layout.addLayout(subject_class_layout, stretch=3)

    def countdown(self):
        dialog = CountdownApp(self,future=QDateTime(2025,2,6,8,0,0))
        dialog.exec()

    def update_version(self):
        if current_version == get_latest_version():
            dialog = custom_dialog_2(self,text=f"Bạn đang dùng phiên bản mới nhất!\n{current_version}",bold=False)
            dialog.exec()
        else:
            string = f'Version hiện tại: {current_version}\nVersion mới nhất: {get_latest_version()}\n\nTải xuống chứ?'
            dialog = update_version_dialog(text=string)
            if dialog.exec():
                dialog.close()
                update_software()

    def sort_subject_list(self):
        data = self.subjects_model._data
        sorted_data = sorted(data, key=lambda subject: class_info_dict[subject] is None)
        self.subjects_model._data = sorted_data
        self.subjects_model.layoutChanged.emit()

    def update_by_faculty_subject(self):
        global subjects_list, class_info_dict, overlap_dict, subject_id_dict
        dialog = subject_by_facaulty_dialog(self)
        if dialog.exec():
            faculty = dialog.get_faculty_name()
            semester = dialog.get_semester()
            data = db.get_subjects_by_faculty_semester(faculty,semester)
            db.delete_class_all()
            for subject_name in data:
                subject_info = (db.get_subject_id_by_name(subject_name),subject_name)
                db.add_subject(subject_info)
            subjects_list = data
            class_info_dict = {subjects_list[i]: None for i in range(len(subjects_list))}
            overlap_dict = {subjects_list[i]: 0 for i in range(len(subjects_list))}
            subject_id_dict = {subject[1]:subject[0] for subject in db.get_subjects_id_name()}
            self.update_current_subject()
            custom_dialog_2(self,'Cập nhật thành công!',False)
            self.subjects_model._data = data
            self.subjects_model.layoutChanged.emit()
        else: 
            pass

    def update_current_subject(self):
        for current_subject in subject_id_dict.values():
            data = search.search(current_subject)
            if data:
                data = self.filter_night(data)
                data = self.filter_over(data)
                class_list = []
                if data:
                    subject_id = data[0][1]
                    subject_name = data[0][3]
                    for row in data:
                        class_row = (
                            row[1],
                            row[2],
                            row[3],
                            row[4],
                            row[5],
                            row[6],
                            row[7],
                            row[8]
                        )
                        class_list.append(class_row)
                    db.update_database(subject_id,subject_name,class_list)
        dialog = custom_dialog_2(self,"Cập nhật thành công!", False)
        dialog.exec()

    def filter_night(self,_data=None):
        if not _data:
            data = self.classes_model._data
        else:
            data = _data
        time_2 = datetime.strptime("18:00", "%H:%M").time()
        if data:
                data_filtered = [class_info for class_info in data if datetime.strptime(class_info[6], "%H:%M").time() < time_2]
        if not _data:
            self.classes_model._data = data_filtered
            self.classes_model.layoutChanged.emit()
            self.classes_filter.hideColumn(0)
            self.classes_filter.setColumnWidth(3,150)
            self.classes_filter.setColumnWidth(4,200)
            self.classes_filter.setColumnWidth(5,80)
        else:
            return data_filtered
            
    def filter_over(self,_data=None):
        if not _data:
            data = self.classes_model._data
        else:
            data = _data
        today = datetime.today().date()
        if data:
            data_filtered = [class_info for class_info in data if datetime.strptime(class_info[9], "%d/%m/%Y").date()>today]
        if not _data:
            self.classes_model._data = data_filtered
            self.classes_model.layoutChanged.emit()
            self.classes_filter.hideColumn(0)
            self.classes_filter.setColumnWidth(3,150)
            self.classes_filter.setColumnWidth(4,200)
            self.classes_filter.setColumnWidth(5,80)
        else:
            return data_filtered

    def close_app(self):
        QApplication.quit()

    def save_file(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Class Infos", "", "JSON Files (*.json)")
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as file:
                    json.dump(class_info_dict, file, indent=4)  # Lưu dữ liệu class_info vào file JSON
            except Exception as e:
                print(f"Lỗi lưu file: {e}")

    def open_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "", "", "JSON Files (*.json)")
        global class_info_dict
        global subjects_list
        global overlap_dict
        global subject_id_dict
        print(class_info_dict)
        if file_path:
            try:
                with open(file_path, "r", encoding="utf-8") as file:
                    data = json.load(file)
                    for key,value in data.items():
                        class_info_dict[key] = value
                    #Chuyển list (mặc định của json) thành tuple
                    for key, value in class_info_dict.items():
                        if value:
                            class_info_dict[key] = tuple(value)
                    
                    subject_lack_list = self.check_subject_in_file()
                    if subject_lack_list:
                        dialog = add_subject_to_database_dialog(self,subject_lack_list)
                        if dialog.exec():
                            self.add_subject_to_database(subject_lack_list)
                            subjects_list = db.get_subjects_name()
                            overlap_dict = {subjects_list[i]: 0 for i in range(len(subjects_list))}
                            subject_id_dict = {subject[1]:subject[0] for subject in db.get_subjects_id_name()}
                            self.update_current_subject()
                            self.subjects_model.layoutChanged.emit()
                    
                    self.check_overlap()
                    self.subject_label.setText(f"Môn học:   {self.subject_count()} đã chọn")
                    self.set_schedule_data()
                    self.customize_header()
            except Exception as e:
                print(f"Lỗi mở file: {e}")

    def add_subject_to_database(self, subject_list=None):
        if subject_list:
            for subject_name in subject_list:
                subject_id = db.get_subject_id_by_name(subject_name)
                if subject_id is not None:  # Kiểm tra nếu subject_id hợp lệ
                    subject_info = (subject_id, subject_name)
                    db.add_subject(subject_info)
                    self.subjects_model._data.append((subject_name))
                    self.subjects_model.layoutChanged.emit()
                else:
                    print(f"Subject '{subject_name}' not found in database!")                 
    
    def check_subject_in_file(self):
        data = []
        for key in class_info_dict.keys():
            if not key in subjects_list:
                data.append(key)
        return data
    
    def on_filter_list_change(self):
        category_text = self.filter_category.currentText()
        text = self.filter_list.currentText()
        subject_name = self.subjects_model.data(
                self.subjects.selectedIndexes()[0], Qt.DisplayRole
            )
        
        if category_text == 'Tất cả':
            self.classes_list = db.get_classes(subject_name)
        elif category_text == 'Giảng viên':
            if text:
                self.classes_list = db.get_classes_by_subject_teacher(subject_name,text)
            else:
                pass
        elif category_text == 'Thứ':
            if text:
                self.classes_list = db.get_classes_by_subject_weekday(subject_name,weekday_int[text])
            else:
                pass
        elif category_text == 'Buổi':
            self.classes_list = db.get_classes_by_subject_phase(subject_name,text)
    
        self.classes_model._data = self.classes_list
        
        #Gọi hàm chọn lớp hiện tại
        self.selected_class_in_subject(subject_name)

        self.classes_model.layoutChanged.emit()
        self.classes.hideColumn(0)

    def on_filter_category_change(self):
        category = self.filter_category.currentText()
        subject_name = self.subjects_model._data[self.subjects.selectedIndexes()[0].row()]
        if category == "Tất cả":
            self.filter_list.clear()
        elif category == "Giảng viên":
            teachers = [item[0] for item in db.get_teachers(subject_name)]
            self.filter_list.clear()
            self.filter_list.addItems(teachers)
        elif category == "Thứ":
            weekdays = [int_weekday[item[0]] for item in db.get_weekdays(subject_name)]
            self.filter_list.clear()
            self.filter_list.addItems(weekdays)
        elif category == "Buổi":
            self.filter_list.clear()
            self.filter_list.addItems(["Sáng", "Chiều"])    
           
    def filter_class(self):
        dialog = custom_dialog_2(self, "Chức năng đang được phát triển!", False)
        dialog.exec()
    
    def refresh_view(self):
        subjects_list = db.get_subjects_name()
        global class_info_dict
        global overlap_dict
        class_info_dict = {subjects_list[i]: None for i in range(len(subjects_list))}
        overlap_dict = {subjects_list[i]: 0 for i in range(len(subjects_list))}
        self.subjects_model.refresh(subjects_list)
        self.set_schedule_data()
        self.subject_label.setText(f"Môn học:   0 đã chọn")
        self.classes.clearSelection()

    def deselect(self):
        self.classes.clearSelection()
        selected_indexes = self.subjects.selectedIndexes()
        if selected_indexes:
            selected_text = self.subjects_model.data(
                selected_indexes[0], Qt.DisplayRole
            )
            class_info_dict[selected_text] = None
            self.check_overlap()
        self.subject_label.setText(f"Môn học:   {self.subject_count()} đã chọn")
        self.set_schedule_data()

    def subject_count(self):
        count = sum(1 for value in class_info_dict.values() if value is not None)
        return count

    def reset_table_span(self):
        row_count = self.table.rowCount()
        column_count = self.table.columnCount()
        self.table.setStyleSheet(
            """
            QTableWidget::item {
                line-height: 15;
            }
            QTableWidget::item:selected {
                background-color:rgb(240, 240, 240); /* Màu nền khi cell được chọn */
                color: black;
            }
            QTableWidget::item:disabled {
                background-color:rgb(242, 242, 242); /* Màu nền khi cell được chọn */
                color: black;
                border-radius: 0px;
            }
                                 """
        )
        for row in range(row_count):
            for col in range(column_count):
                if self.table.rowSpan(row, col) > 1:
                    self.table.setSpan(row, col, 1, 1)

    def delete_class(self):
        if self.classes.selectedIndexes():
            indexes = self.classes.selectedIndexes()
            index = indexes[0]
            class_info = self.classes_model._data[index.row()]
            class_name = class_info[3]
            class_weekday = int_weekday[class_info[5]]
            class_time_start = class_info[6]
            class_time_end = class_info[7]
            class_id = class_info[2]
            dialog = custom_dialog(
                self,
                f"Bạn muốn xóa lớp\n【 {class_name} 】\n{class_weekday}\n{class_time_start} ➞ {class_time_end}\nMã lớp:  {class_id}?",
                False,
            )
            if dialog.exec():
                try:
                    db.delete_class(class_id)
                    self.on_subject_view_change()
                except:
                    pass
        else:
            dialog = custom_dialog_2(self, "Vui lòng chọn lớp cần xóa!", False)
            dialog.exec()

    def delete_subject(self):
        if self.subjects.selectedIndexes():
            indexes = self.subjects.selectedIndexes()
            index = indexes[0]
            subject_name = self.subjects_model._data[index.row()]
            dialog = custom_dialog(self, f"Bạn muốn xóa môn 【 {subject_name} 】?")
            if dialog.exec():
                try:
                    db.delete_subject(subject_name)
                    del self.subjects_model._data[index.row()]
                    self.subjects_model.dataChanged.emit(index, index)
                except:
                    pass
        else:
            dialog = custom_dialog_2(self, "Vui lòng chọn môn học cần xóa!", False)
            dialog.exec()

    def delete_all(self):
        global subjects_list
        dialog = custom_dialog(
            self,
            "Bạn muốn xóa TẤT CẢ dữ liệu?\n\nLưu ý: Không thể hoàn tác!\n",
            bold=False,
        )
        if dialog.exec():
            try:
                db.delete_class_all()
                self.subjects_model._data = []
                subjects_list = []
                self.classes_model._data = []
                self.classes_model.layoutChanged.emit()
                self.subjects_model.layoutChanged.emit()
            except:
                pass

    def add_subject(self):
        dialog = find_subject_form(subject_list_model=self.subjects_model)
        global subjects_list, class_info_dict, overlap_dict, subject_id_dict
        old_subjects_list_len = len(subjects_list)
        if dialog.exec() == 0:
            subjects_list = db.get_subjects_name()
            if len(subjects_list) > old_subjects_list_len:
                self.subjects_model._data = subjects_list
                for subject in subjects_list:
                    if class_info_dict.get(subject) is None:
                        class_info_dict[subject] = None
                    if overlap_dict.get(subject) is None:
                        overlap_dict[subject] = 0
                    if subject_id_dict.get(subject) is None:
                        subject_id_dict[subject] = db.get_subject_id_by_name(subject)
                self.update_current_subject()
                self.subjects_model.layoutChanged.emit()

    def on_subject_view_change(self):
        selected_indexes = self.subjects.selectedIndexes()
        self.filter_label.setVisible(True)
        self.filter_category.setVisible(True)
        self.filter_list.setVisible(True)
        self.filter_category.setCurrentIndex(0)
        if selected_indexes:
            selected_text = self.subjects_model.data(
                selected_indexes[0], Qt.DisplayRole
            )

            self.classes_list = db.get_classes(selected_text)
            self.classes_model._data = self.classes_list
            self.class_label.setText(f"Lớp học tương ứng:    【 {selected_text} 】")
            self.classes.clearSelection()
                
            #Gọi hàm chọn lớp hiện tại
            self.selected_class_in_subject(selected_text)

            self.classes_model.layoutChanged.emit()
            self.set_schedule_data()
            self.classes.hideColumn(0)
            self.on_filter_category_change()
    
    def selected_class_in_subject(self,subject_name):
        class_info = class_info_dict[subject_name]
        model = self.classes_model
        row_count = model.rowCount()
        match_index = -1
        if class_info:
            class_id = class_info[2]
            weekday = class_info[5]
            time_start = class_info[6]
            for row in range(row_count):
                check = (class_id == model.index(row, 2).data() and weekday == model.index(row, 5).data() and time_start == model.index(row, 6).data())
                if check:
                    self.classes.selectRow(row)
                    match_index = row
                    break
            if match_index == -1:
                self.classes.clearSelection()

    def on_class_table_change(self):
        selected_index = self.classes.selectionModel().selectedIndexes()
        # Hiển thị nút Bỏ Chọn khi có lớp được chọn
        if selected_index:
            self.deselect_button.setDisabled(False)
            self.deselect_button.setStyleSheet(
                """
                QPushButton {
                    background-color: #4686e3;
                    color: white;
                    border-radius: 6px;
                    outline: none;
                    font-weight: bold;
                    font-family: -apple-system, BlinkMacSystemFont, 'Roboto', sans-serif;
                    border-bottom: 3px solid #3f73bf;
                }
                QPushButton:hover {
                    background-color: #4b82d1;
                    outline: None;
                }
                                          """
            )
        else:
            self.deselect_button.setDisabled(True)
            self.deselect_button.setStyleSheet(self.disable_button_style_string)
        # Hết phần hiển thị nút Bỏ Chọn
        if selected_index:
            global class_info_dict
            global overlap_dict
            selected_row = list(set(index.row() for index in selected_index))
            row = selected_row[-1]
            row_data = []
            class_info = ()
            for col in range(self.classes_model.columnCount()):
                index = self.classes_model.index(row, col)
                row_data.append(self.classes_model.data(index))
                class_info += (self.classes_model.data(index),)
            class_info_dict[row_data[3]] = class_info
            self.check_overlap()
            
            self.subject_label.setText(f"Môn học:   {self.subject_count()} đã chọn")
            self.set_schedule_data()
    
    def set_schedule_data(self):

        self.table.clearContents()
        self.reset_table_span()
        
        time_rows = {
            "07:00": 1,
            "07:45": 2,
            "08:30": 3,
            "08:55": 3,
            "09:40": 4,
            "10:30": 5,
            "11:15": 6,
            "12:00": 7,
            "12:45": 8,
            "13:30": 9,
            "14:15": 10,
            "15:00": 11,
            "15:25": 11,
            "16:10": 12,
            "16:55": 13,
            "17:40": 14,
        }
        
        for subject, info in class_info_dict.items():
            if info:
                for entry in ([info] if isinstance(info, tuple) else info):  # Xử lý cả trường hợp tuple hoặc list
                    (
                        id,
                        subject_id,
                        class_id,
                        class_name,
                        teacher,
                        weekday,
                        time_start,
                        time_end,
                        room,
                    ) = entry
                    start_row = time_rows[time_start]
                    end_row = time_rows[time_end] - 1
                    column = weekday_int[weekday] - 2

                    color = "#ebf5f1" if column % 2 == 0 else "#ebf0ee"
                    for row in range(start_row, end_row + 1):
                        if end_row - start_row + 1 > 2:
                            cell_name = QLabel(
                                f"""
                                <html>
                                    <body style="line-height: 20px;">
                                        《 {class_id} 》<br>
                                        <span style="font-weight:bold; color: #096141;">{subject}</span><br>
                                        GV: {teacher}<br>
                                        <span style="font-weight:bold;">{time_start} ➞ {time_end}</span>
                                    </body>
                                </html>
                                """
                            )
                        else:
                            cell_name = QLabel(
                                f"""
                                <html>
                                    <body style="line-height: 10px;">
                                        《 {class_id} 》<br>
                                        <span style="font-weight:bold; color: #096141;">{subject}</span><br>
                                        <span style="font-weight:bold;">{time_start} ➞ {time_end}</span>
                                    </body>
                                </html>
                                """
                            )
                        cell_name.setAlignment(Qt.AlignCenter)
                        cell_name.setStyleSheet(
                            f"""
                            QLabel {{
                                color: #333333;
                                font-size: 12px;
                                background-color: {color};
                            }}
                            """
                        )
                        self.table.setCellWidget(row, column, cell_name)
                    self.table.setSpan(start_row, column, end_row - start_row + 1, 1)

        # Tạo dòng "Chiều"
        self.table.setSpan(7, 0, 1, 7)
        item = self.table.item(7, 0)
        if item is None:
            item = QTableWidgetItem("Chiều")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
            self.table.setItem(7, 0, item)

        # Tạo dòng "Sáng"
        self.table.setSpan(0, 0, 1, 7)
        item = self.table.item(0, 0)
        if item is None:
            item = QTableWidgetItem("Sáng")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
            self.table.setItem(0, 0, item)

    # Hàm chạy khi nhấn vào Lưu
    def export_schedule(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,  # Widget cha
            "Lưu tệp",  # Tiêu đề hộp thoại
            "Output",  # Đường dẫn mặc định
            "PNG Files (*.png)",  # Bộ lọc định dạng tệp
        )
        if file_path:
            print(f"Đường dẫn đã chọn: {file_path}")
            export_to_image(self.table, file_path)

    def closeEvent(self, event):
        dialog = custom_dialog(self, "Bạn chắc chắn muốn thoát chứ?")
        dialog.show()
        if dialog.exec():
            event.accept()
        else:
            event.ignore()

    def customize_header(self):
        """Tùy chỉnh style cho header"""
        # Đặt căn lề chữ trong header (nằm sát cạnh dưới)
        self.table.horizontalHeader().setDefaultAlignment(
            Qt.AlignBottom | Qt.AlignHCenter
        )

        self.table.verticalHeader().setStyleSheet(
            """
            QHeaderView::section {
                color: #333333;             /* Màu chữ */
                font-size: 12px;           /* Kích thước chữ */
                padding-left: 25px;  /* Đẩy chữ sát phải */
                padding-right: 10px;  /* Đẩy chữ sát phải */
                padding-top: 20px;  /* Đẩy chữ xuống dưới */
                    }
                """
        )

    #Hàm kiểm tra trùng lịch, chạy khi chọn lớp học
    def check_overlap(self):
        global overlap_dict
        overlap_dict = {subjects_list[i]: 0 for i in range(len(subjects_list))}
        for subject1, values1 in class_info_dict.items():
            if values1:
                for subject2, values2 in class_info_dict.items():
                    if values2 is None or subject1 == subject2:
                        continue

                    # Kiểm tra nếu hai môn không cùng thứ
                    if len(values1) > 6 and len(values2) > 6:
                        if not values1[5] == values2[5]:
                            continue

                    # Kiểm tra nếu có trùng thời gian
                    if len(values1) > 6 and len(values2) > 6:
                        check = (values1[6] >= values2[6] and values1[6] < values2[7]) or (
                            values1[7] > values2[6] and values1[7] <= values2[7]
                        )
                        if check:
                            overlap_dict[subject1] = 1
                            overlap_dict[subject2] = 1

        self.subjects_model.dataChanged.emit(
            self.subjects_model.index(0),
            self.subjects_model.index(len(self.subjects_model._data) - 1),
            [Qt.DisplayRole, Qt.ForegroundRole, Qt.FontRole],
        )                                   
                            
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(
        """
        ScheduleWindow, QDialog {
            background-color: rgb(255, 255, 255);
        }
        QPushButton {
            background-color: #4686e3;
            color: white;
            border-radius: 6px;
            outline: none;
            font-weight: bold;
            font-family: -apple-system, BlinkMacSystemFont, 'Roboto', sans-serif;
            border-bottom: 3px solid #3f73bf;
        }
        QPushButton:hover {
            background-color: #4b82d1;
            outline: None;
        }
        QLineEdit {
                padding: 2px 5px;
                border-radius: 0px;
                border: 2px solid #f0f2f1;
        }                
                      """
    )
    window = ScheduleWindow()
    window.showMaximized()
    sys.exit(app.exec())
